# Fuzzy Logic System untuk Memilih 5 Restoran Terbaik
# Nama: Maria Nathasya Desfera Pangestu (2211104008)
# Nama: Lintang Suminar Tyas Wening (2211104009)
# Kelas: SE0601
# Mata Kuliah: Kecerdasan Buatan

import openpyxl

# Fungsi Membership

def membership_servis(servis):
    """
    Menghitung derajat keanggotaan untuk kategori servis (buruk, sedang, bagus).
    Menggunakan bentuk segitiga untuk keanggotaan fuzzy.
    """
    # Buruk: 0-50 (semakin kecil semakin buruk)
    if servis <= 25:
        buruk = 1
    elif 25 < servis <= 50:
        buruk = (50 - servis) / 25
    else:
        buruk = 0

    # Sedang: 40-70 (naik dari 40 ke 55, turun dari 55 ke 70)
    if 40 < servis <= 55:
        sedang = (servis - 40) / 15
    elif 55 < servis <= 70:
        sedang = (70 - servis) / 15
    else:
        sedang = 0

    # Bagus: 60-100 (naik dari 60 ke 80, lalu penuh)
    if 60 < servis <= 80:
        bagus = (servis - 60) / 20
    elif 80 < servis <= 100:
        bagus = 1
    else:
        bagus = 0

    return {'buruk': max(min(buruk, 1), 0),
            'sedang': max(min(sedang, 1), 0),
            'bagus': max(min(bagus, 1), 0)}

def membership_harga(harga):
    """
    Menghitung derajat keanggotaan untuk kategori harga (murah, sedang, mahal).
    Menggunakan fungsi segitiga.
    
    Batas harga:
    - Murah: 0 - 30,000
    - Sedang: 25,000 - 45,000
    - Mahal: 40,000 - 70,000
    """
    if harga <= 20000:
        murah = 1
    elif 20000 < harga <= 30000:
        murah = (30000 - harga) / 10000
    else:
        murah = 0

    if 25000 < harga <= 35000:
        sedang = (harga - 25000) / 10000
    elif 35000 < harga <= 45000:
        sedang = (45000 - harga) / 10000
    else:
        sedang = 0

    if 40000 < harga <= 55000:
        mahal = (harga - 40000) / 15000
    elif 55000 < harga <= 70000:
        mahal = 1
    else:
        mahal = 0

    return {'murah': max(min(murah, 1), 0),
            'sedang': max(min(sedang, 1), 0),
            'mahal': max(min(mahal, 1), 0)}

# Fungsi Inferensi

def inferensi(mb_servis, mb_harga):
    rules = []

    rules.append(('sangat_baik', min(mb_servis['bagus'], mb_harga['murah'])))
    rules.append(('biasa_saja', min(mb_servis['bagus'], mb_harga['sedang'])))
    rules.append(('biasa_saja', min(mb_servis['bagus'], mb_harga['mahal'])))

    rules.append(('biasa_saja', min(mb_servis['sedang'], mb_harga['murah'])))
    rules.append(('biasa_saja', min(mb_servis['sedang'], mb_harga['sedang'])))
    rules.append(('buruk', min(mb_servis['sedang'], mb_harga['mahal'])))

    rules.append(('buruk', min(mb_servis['buruk'], mb_harga['murah'])))
    rules.append(('buruk', min(mb_servis['buruk'], mb_harga['sedang'])))
    rules.append(('buruk', min(mb_servis['buruk'], mb_harga['mahal'])))

    hasil = {'buruk': 0, 'biasa_saja': 0, 'sangat_baik': 0}
    for kategori, nilai in rules:
        if nilai > hasil[kategori]:
            hasil[kategori] = nilai
    return hasil

# Fungsi Defuzzifikasi

def defuzzifikasi(hasil_inferensi):
    """
    Melakukan proses defuzzifikasi menggunakan metode rata-rata terbobot (weighted average).

    Mapping nilai (mengikuti label pada inferensi):
    - 'buruk'        : 25
    - 'biasa_saja'   : 50
    - 'sangat_baik'  : 85
    """
    bobot = {
        'buruk': 25,
        'biasa_saja': 50,
        'sangat_baik': 85
    }

    nilai_total = (hasil_inferensi['buruk'] * bobot['buruk'] +
                   hasil_inferensi['biasa_saja'] * bobot['biasa_saja'] +
                   hasil_inferensi['sangat_baik'] * bobot['sangat_baik'])

    total_keanggotaan = (hasil_inferensi['buruk'] +
                         hasil_inferensi['biasa_saja'] +
                         hasil_inferensi['sangat_baik'])

    if total_keanggotaan == 0:
        return 0

    return nilai_total / total_keanggotaan

# Fungsi Menentukan Keterangan Kualitas

def keterangan_kualitas(skor):
    if skor < 35:
        return "Buruk"
    elif skor < 65:
        return "Biasa Saja"
    else:
        return "Sangat Baik"

# Membaca data dari restoran.xlsx

def baca_data(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        id_restoran, servis, harga = row
        data.append({'id': id_restoran, 'pelayanan': servis, 'harga': harga})
    return data

# Menyimpan output ke peringkat.xlsx

def simpan_output(filename, hasil):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['ID Restoran', 'Pelayanan', 'Harga', 'Skor Kelayakan', 'Keterangan Kualitas'])
    for item in hasil:
        ws.append([item['id'], item['pelayanan'], item['harga'], item['skor'], item['keterangan']])
    wb.save(filename)

# Main Program

def main():
    data = baca_data('restoran.xlsx')
    hasil = []

    for restoran in data:
        mb_servis = membership_servis(restoran['pelayanan'])
        mb_harga = membership_harga(restoran['harga'])
        hasil_inferensi = inferensi(mb_servis, mb_harga)
        skor = defuzzifikasi(hasil_inferensi)
        ket = keterangan_kualitas(skor)
        hasil.append({'id': restoran['id'], 'pelayanan': restoran['pelayanan'],
                      'harga': restoran['harga'], 'skor': skor, 'keterangan': ket})

    hasil = sorted(hasil, key=lambda x: x['skor'], reverse=True)
    top5 = hasil[:5]

    simpan_output('peringkat.xlsx', top5)

    print("5 Restoran Terbaik:")
    for item in top5:
        print(f"ID: {item['id']}, Pelayanan: {item['pelayanan']}, Harga: {item['harga']}, Skor: {item['skor']:.2f}, Kualitas: {item['keterangan']}")

if __name__ == "__main__":
    main()
